"""Excel export for a client's full measurement history."""
from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import MeasurementSet

# Order matters — these become the column headers.
COLUMNS: list[tuple[str, str]] = [
    ('Visit', 'created_at'),
    ('Garment types', 'garment_types'),
    ('Count', 'garment_count'),
    # Upper
    ('U · Length', 'upper_length'),
    ('U · Shoulder', 'upper_shoulder'),
    ('U · Sleeve', 'upper_sleeve'),
    ('U · ½ Sleeve', 'upper_half_sleeve'),
    ('U · Chest', 'upper_chest'),
    ('U · Waist', 'upper_waist'),
    ('U · Hip', 'upper_hip'),
    ('U · Cuff', 'upper_cuff'),
    ('U · Collar', 'upper_collar'),
    ('U · Arms', 'upper_arms'),
    # Lower
    ('L · Length', 'lower_length'),
    ('L · Bottom', 'lower_bottom'),
    ('L · Knee', 'lower_knee'),
    ('L · Waist', 'lower_waist'),
    ('L · Hip', 'lower_hip'),
    ('L · Seat round', 'lower_seat_round'),
    ('L · Inseam', 'lower_inseam'),
    ('L · Thigh', 'lower_thigh'),
    # Suit
    ('Suit · Lapel', 'suit_lapel_style'),
    ('Suit · Buttons', 'suit_button_stance'),
    ('Suit · Vent', 'suit_vent'),
    # Notes
    ('Fabric', 'fabric_details'),
    ('Customization', 'customization_notes'),
]


def _gold_header_fill() -> PatternFill:
    return PatternFill('solid', fgColor='FFCBA624')


def _navy_text() -> Font:
    return Font(name='Calibri', bold=True, color='FF1D1840')


def _row_value(ms: MeasurementSet, attr: str):
    if attr == 'created_at':
        return ms.created_at.strftime('%Y-%m-%d %H:%M')
    if attr == 'garment_types':
        return ', '.join(ms.garment_types or [])
    return getattr(ms, attr) or ''


def build_measurement_workbook(client, sets: Iterable[MeasurementSet]) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Measurements'

    # Title block
    ws['A1'] = 'VERSUITALITY · Bespoke Tailoring'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='FF261F53')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws['A2'] = f'Client: {client.full_name}  ·  ID: {client.client_id}  ·  Mobile: {client.mobile}'
    ws['A2'].font = Font(name='Calibri', italic=True, color='FF52515A')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    header_row = 4
    thin = Side(border_style='thin', color='FFE0DACA')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for idx, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.font = _navy_text()
        cell.fill = _gold_header_fill()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for r, ms in enumerate(sets, start=header_row + 1):
        for c, (_, attr) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=_row_value(ms, attr))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            cell.font = Font(name='Calibri', size=10)

    for col in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions['A'].width = 18  # visit
    ws.column_dimensions['B'].width = 24  # garment types

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    filename = f'measurements_{client.client_id}.xlsx'
    return buf.getvalue(), filename
