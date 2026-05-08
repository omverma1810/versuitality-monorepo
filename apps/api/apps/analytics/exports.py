"""Date-range Excel export of orders for offline analysis."""
from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.orders.models import Order

GOLD_FILL = PatternFill('solid', fgColor='FFCBA624')
NAVY_FONT = Font(name='Calibri', bold=True, color='FF1D1840')
BODY_FONT = Font(name='Calibri', size=10)
THIN = Side(border_style='thin', color='FFE0DACA')
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


COLUMNS: list[tuple[str, callable]] = [
    ('Order ID', lambda o: o.order_id),
    ('Created', lambda o: o.created_at.strftime('%Y-%m-%d %H:%M')),
    ('Status', lambda o: o.get_status_display()),
    ('Order type', lambda o: o.get_order_type_display()),
    ('Client ID', lambda o: o.client.client_id),
    ('Client name', lambda o: o.client.full_name),
    ('Mobile', lambda o: o.client.mobile),
    ('Email', lambda o: o.client.email),
    ('Garments', lambda o: o.garment_summary),
    ('Subtotal', lambda o: float(o.subtotal or 0)),
    ('Advance', lambda o: float(o.advance or 0)),
    ('Balance', lambda o: float(o.balance or 0)),
    ('Trial date', lambda o: o.trial_date.isoformat() if o.trial_date else ''),
    ('Delivery date', lambda o: o.delivery_date.isoformat() if o.delivery_date else ''),
    ('Delivered at', lambda o: o.delivered_at.strftime('%Y-%m-%d %H:%M') if o.delivered_at else ''),
    ('Created by', lambda o: o.created_by.full_name if o.created_by_id else ''),
]


def build_orders_workbook(frm: date, to: date) -> tuple[bytes, str]:
    qs = (
        Order.objects.select_related('client', 'created_by')
        .prefetch_related('line_items')
        .filter(created_at__date__gte=frm, created_at__date__lte=to)
        .order_by('-created_at')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Orders'

    ws['A1'] = 'VERSUITALITY · Order export'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='FF261F53')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    ws['A2'] = f'Range: {frm.isoformat()} to {to.isoformat()}  ·  Total orders: {qs.count()}'
    ws['A2'].font = Font(name='Calibri', italic=True, color='FF52515A')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    header_row = 4
    for idx, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=idx, value=label)
        cell.font = NAVY_FONT
        cell.fill = GOLD_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    for r, order in enumerate(qs, start=header_row + 1):
        for c, (_, accessor) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=accessor(order))
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.font = BODY_FONT
            cell.border = BORDER

    widths = [16, 18, 22, 16, 14, 26, 14, 28, 36, 12, 12, 12, 12, 12, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    filename = f'versuitality_orders_{frm.isoformat()}_to_{to.isoformat()}.xlsx'
    return buf.getvalue(), filename
